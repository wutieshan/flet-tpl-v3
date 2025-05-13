# reference: https://openpyxl-chinese-docs.readthedocs.io/zh-cn/latest/tutorial.html
from openpyxl import Workbook, load_workbook
from openpyxl.cell import Cell
from openpyxl.drawing.image import Image
from openpyxl.worksheet.worksheet import Worksheet

from app.utils.log_util import log


class Xlsx:
    def __init__(self, filepath: str):
        self.filepath = filepath
        self.workbook: Workbook = None

    def open(self) -> Workbook:
        if self.workbook is not None:
            log.info(f"[silent] workbook already opened: {self.filepath}")
            return self.workbook

        try:
            self.workbook = load_workbook(self.filepath)
        except FileNotFoundError:
            self.workbook = Workbook()
            self.save()

        return self.workbook

    def __enter__(self):
        self.open()
        return self

    def __exit__(self, exc_type, exc_val, exc_tb) -> None:
        self.close()

    def close(self) -> None:
        if self.workbook is None:
            return

        self.workbook.close()

    def save(self) -> None:
        self.workbook.save(self.filepath)

    def save_as(self, filepath: str) -> None:
        self.workbook.save(filepath)

    def create_sheet(self, title: str, index: int | None = None) -> None:
        if title in self.workbook.sheetnames:
            log.info(f"[silent] sheet already exists: {title}")
            return

        if index is None:
            self.workbook.create_sheet(title)
        else:
            self.workbook.create_sheet(title, index)

    def rename_sheet(self, title: str, sheet: str | int | None = None):
        if sheet is None:
            self.workbook.active.title = title
        elif isinstance(sheet, int) and 0 <= sheet < len(self.workbook.worksheets):
            self.workbook.worksheets[sheet].title = title
        elif isinstance(sheet, str) and sheet in self.workbook.sheetnames:
            self.workbook[sheet].title = title
        else:
            log.warning(f"sheet not found: {sheet}, do nothing")

    def get_sheet(self, sheet: str | int | None = None) -> Worksheet:
        if isinstance(sheet, int) and 0 <= sheet < len(self.workbook.worksheets):
            return self.workbook.worksheets[sheet]
        elif isinstance(sheet, str) and sheet in self.workbook.sheetnames:
            return self.workbook[sheet]
        else:
            return self.workbook.active

    def delete_sheet(self, sheet: str | int | None = None) -> None:
        if len(self.workbook.worksheets) == 1:
            log.warning("cannot delete the only one sheet, do nothing")
            return

        if isinstance(sheet, int) and 0 <= sheet < len(self.workbook.worksheets):
            self.workbook.remove(self.workbook.worksheets[sheet])
        elif isinstance(sheet, str) and sheet in self.workbook.sheetnames:
            self.workbook.remove(self.workbook[sheet])
        else:
            # self.workbook.remove(self.workbook.active)
            pass

    def switch_sheet(self, sheet: str | int | None = None):
        if isinstance(sheet, int) and 0 <= sheet < len(self.workbook.worksheets):
            self.workbook.active = sheet
        elif isinstance(sheet, str) and sheet in self.workbook.sheetnames:
            self.workbook[sheet].activate()
        else:
            log.warning(f"sheet not found: {sheet=}, do nothing")

    @staticmethod
    def col_str2int(label: str) -> int:
        assert label.isalpha(), "column label must be alphabetic"
        return sum((ord(v) - 64) * 26**i for i, v in enumerate(label.upper()[::-1]))

    @staticmethod
    def col_int2str(col: int) -> str:
        """
        get_column_letter(col)
        """

        if col < 1:
            raise ValueError("column index must greater then 0")

        label = ""
        while col > 0:
            col, remainder = divmod(col - 1, 26)
            label = chr(ord("A") + remainder) + label

        return label

    def get_cell(self, row: int, col: int | str, sheet: str | int | None = None) -> Cell:
        ws = self.get_sheet(sheet)

        if isinstance(col, str):
            col = self.col_str2int(col)

        return ws.cell(row=row, column=col)

    def add_image(self, path: str, row: int, col: int | str, sheet: str | int | None = None, scaling: float = 1.0) -> None:
        img = Image(path)
        img.width = int(img.width * scaling)
        img.height = int(img.height * scaling)
        ws = self.get_sheet(sheet)

        if isinstance(col, int):
            col = self.col_int2str(col)

        ws.add_image(img, f"{col}{row}")

    def delete_all_images(self, sheet: str | int | None = None) -> None:
        ws = self.get_sheet(sheet)
        ws._images.clear()
